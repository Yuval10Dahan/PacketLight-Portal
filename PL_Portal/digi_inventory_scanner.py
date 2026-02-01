import socket
import ipaddress
import telnetlib
import re
import time
from typing import List, Dict, Any
import openpyxl # Import the openpyxl library

# --- Global Settings ---
# Use a list of tuples to define multiple IP ranges.
# A range with "0" as the start or end IP will be skipped.
IP_RANGES = [
    ("172.16.10.1", "172.16.10.3"),
    ("172.16.20.1", "172.16.20.3"),
    ("172.16.30.1", "172.16.30.3"),
    ("172.16.40.1", "172.16.40.3"),
    ("10.30.5.0", "10.30.5.10")
]

PORTS_FOR_DISCOVERY = list(range(2016, 2018))
PORTS_FOR_TELNET_16 = list(range(2001, 2017))
PORTS_FOR_TELNET_32 = list(range(2001, 2033))

# --- Credentials for Telnet ---
TELNET_USERNAME = "tech"
TELNET_PASSWORD = "packetlight"
# ------------------------------

def check_ports(ip_address: str, ports: List[int]) -> List[int]:
    """
    Checks which ports from a given list are open on a specified IP address.
    """
    open_ports = []
    for port in ports:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(0.5)
        
        try:
            if sock.connect_ex((ip_address, port)) == 0:
                open_ports.append(port)
        except (socket.gaierror, socket.error):
            pass
        finally:
            sock.close()
            
    return open_ports

def scan_and_process_results(start_ip: str, end_ip: str, ports_to_check: List[int]) -> List[Dict[str, Any]]:
    """
    Scans a given IP range for open ports and returns a list of dictionaries.
    """
    results_array = []
    
    try:
        start_ip_obj = ipaddress.IPv4Address(start_ip)
        end_ip_obj = ipaddress.IPv4Address(end_ip)
    except ipaddress.AddressValueError as e:
        print(f"Error with IP address format: {e}")
        return results_array
    
    current_ip = int(start_ip_obj)
    while current_ip <= int(end_ip_obj):
        ip_str = str(ipaddress.IPv4Address(current_ip))
        print(f"Scanning {ip_str}...")
        
        open_ports_for_discovery = check_ports(ip_str, ports_to_check)
        
        if open_ports_for_discovery:
            num_ports_reported = 0
            if 2017 in open_ports_for_discovery:
                num_ports_reported = 32
            elif 2016 in open_ports_for_discovery:
                num_ports_reported = 16
            
            if num_ports_reported > 0:
                results_array.append({
                    "ip": ip_str,
                    "open_ports_count": num_ports_reported
                })
        
        current_ip += 1
    
    return results_array

def connect_via_telnet(scan_results: List[Dict[str, Any]], username: str, password: str) -> List[Dict[str, str]]:
    """
    Connects via Telnet to the IPs, authenticates, runs commands, and extracts data.
    """
    extracted_data = []
    
    print("\n--- Starting Telnet Connections and Data Extraction ---")
    
    for result in scan_results:
        ip_address = result["ip"]
        port_count = result["open_ports_count"]
        
        ports_to_connect = []
        if port_count == 32:
            ports_to_connect = PORTS_FOR_TELNET_32
        elif port_count == 16:
            ports_to_connect = PORTS_FOR_TELNET_16
        print(f"Attempting to connect to {ip_address}  ports: {ports_to_connect}")
        for port in ports_to_connect:
            tn = None
            try:
                tn = telnetlib.Telnet(ip_address, port, timeout=5)
                

                tn.write(b"/\n")
                tn.read_until(b">>", timeout=5)
                
                tn.write(b"login\n")
                
                tn.read_until(b"User: ", timeout=5)
                tn.write(username.encode('ascii') + b"\n")
                
                tn.read_until(b"Password: ", timeout=5)
                tn.write(password.encode('ascii') + b"\n")

                tn.read_until(b">>", timeout=5)

                tn.write(b"/\n")

                # Capture the full prompt after login
                full_prompt = tn.read_until(b">>", timeout=5).decode('ascii', errors='ignore')
                
                # Regex to extract the product name from the prompt
                product_match = re.search(r"([A-Z0-9-]+):", full_prompt)
                product_name = product_match.group(1) if product_match else "None"

                tn.write(b"c i eth ip\n")
                output_ip = tn.read_until(b">>", timeout=5).decode('ascii')
                
                ip_match = re.search(r"Addr is\s+(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})", output_ip)
                received_ip = ip_match.group(1) if ip_match else "None"
                                      
                extracted_data.append({
                    "connected_ip": ip_address,
                    "connected_port": str(port),
                    "received_ip": received_ip,
                    "product_name": product_name
                })
                
                # Print the result immediately after extracting the data
                print(f"Connected to Digi IP: {ip_address}:{port} | Device: {product_name} : IP {received_ip}")
                
            except telnetlib.socket.timeout:
                print(f"Connection to {ip_address}:{port} timed out.")
            except Exception as e:
                print(f"Failed to connect or extract data from {ip_address}:{port}. Error: {e}")
            finally:
                if tn:
                    tn.close()
            
    
    return extracted_data

# --- New function to export data to an Excel file ---
def export_to_excel(final_data: List[Dict[str, str]], filename="telnet_scan_results.xlsx"):
    """
    Exports the collected data to an Excel file, formatted by IP address.
    
    Args:
        final_data (List[Dict[str, str]]): The list of dictionaries with the extracted data.
        filename (str): The name of the Excel file to save.
    """
    if not final_data:
        print("No data to export. Skipping Excel file creation.")
        return
        
    print(f"\n--- Exporting data to {filename} ---")
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Scan Results"
    
    current_row = 1
    
    # Group data by the connected IP address
    grouped_data = {}
    for entry in final_data:
        ip = entry["connected_ip"]
        if ip not in grouped_data:
            grouped_data[ip] = []
        grouped_data[ip].append(entry)
        
    # Write the grouped data to the sheet
    for ip, entries in grouped_data.items():
        # Add the main IP as a header
        sheet.cell(row=current_row, column=1, value=f"IP: {ip}")
        current_row += 1
        
        # Add a sub-header for the data
        sheet.cell(row=current_row, column=1, value="Port")
        sheet.cell(row=current_row, column=2, value="Received IP")
        sheet.cell(row=current_row, column=3, value="Product Name")
        current_row += 1
        
        # Write the data for each port
        for entry in entries:
            sheet.cell(row=current_row, column=1, value=entry["connected_port"])
            sheet.cell(row=current_row, column=2, value=entry["received_ip"])
            sheet.cell(row=current_row, column=3, value=entry["product_name"])
            current_row += 1
        
        # Add a blank row for spacing
        current_row += 1
        
    try:
        workbook.save(filename)
        print(f"Data successfully exported to {filename}.")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

def print_results(final_data: List[Dict[str, str]]):
    """
    Prints the collected data in a formatted way.
    """
    print("\n--- Final Results ---")
    if final_data:
        for data in final_data:
            print(f"Connected to: {data['connected_ip']}:{data['connected_port']}")
            print(f"  Received IP: {data['received_ip']}")
            print(f"  Product Name: {data['product_name']}")
            print("-" * 30)
    else:
        print("No data was extracted from any device.")

def main():
    """
    Main function to orchestrate the entire process.
    """
    total_scan_results = []
    
    for start_ip, end_ip in IP_RANGES:
        if start_ip == "0" or end_ip == "0":
            print(f"Skipping scan for range: ({start_ip}, {end_ip})")
            continue
            
        print(f"\n--- Starting scan for range: {start_ip} to {end_ip} ---")
        
        scan_results = scan_and_process_results(start_ip, end_ip, PORTS_FOR_DISCOVERY)
        total_scan_results.extend(scan_results)

    print("\n--- Initial Scan Results Summary ---")
    if total_scan_results:
        for result in total_scan_results:
            print(f"IP: {result['ip']} - Detected Ports Count: {result['open_ports_count']}")
        
        extracted_data = connect_via_telnet(total_scan_results, TELNET_USERNAME, TELNET_PASSWORD)
        
        print_results(extracted_data)
        
        # Call the new function to export the data to Excel
        export_to_excel(extracted_data)
    else:
        print("No devices found with the specified ports open in any of the ranges.")

if __name__ == "__main__":
    main()